"""
cli_tool.py — CSV folder to Excel workbook converter.

Usage:
    python cli_tool.py <input_dir> --output <output.xlsx> [--verbose]

Produces a single Excel workbook with:
  - One sheet per CSV file (alphabetical order)
  - A "Summary" sheet (always first) with row counts, column counts, and column names
  - A bar chart on the Summary sheet showing row counts per file

Handles: BOM-encoded files, empty CSVs, heterogeneous schemas, sheet name collisions,
         Excel row/column limits, and files that cannot be parsed.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path
from typing import Optional, TypedDict

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLS = 16_384
SHEET_NAME_MAX_LEN = 31
LARGE_FILE_WARN_ROWS = 500_000
COL_WIDTH_MIN = 8
COL_WIDTH_MAX = 60
SUMMARY_SHEET_NAME = "Summary"
RESERVED_SUMMARY_STEM = "summary"  # case-insensitive; triggers rename

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="366092")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

log = logging.getLogger(__name__)


class ResultEntry(TypedDict):
    """Typed record for one CSV's processing outcome."""

    file: str
    rows: Optional[int]
    columns: Optional[int]
    column_names: Optional[str]
    error: Optional[str]


# ---------------------------------------------------------------------------
# T1 — CLI / argument parsing
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    """Build and return the argument parser for the CLI."""
    parser = argparse.ArgumentParser(
        prog="cli_tool",
        description=(
            "Convert a folder of CSV files into a single Excel workbook. "
            "Produces one sheet per CSV plus a Summary sheet with row counts and a bar chart."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python cli_tool.py ./data --output report.xlsx\n"
            "  python cli_tool.py ./data --output ./out/report.xlsx --verbose\n"
        ),
    )
    parser.add_argument(
        "input_dir",
        type=Path,
        help="Directory containing CSV files to process.",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        metavar="PATH",
        help="Path for the output Excel workbook (.xlsx). Parent dirs are created if needed.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable DEBUG-level logging.",
    )
    return parser


def validate_args(args: argparse.Namespace) -> None:
    """
    Validate parsed CLI arguments, exiting with code 1 on any error.

    Args:
        args: Parsed namespace from argparse.
    """
    if not args.input_dir.exists():
        _exit_error(f"Input directory does not exist: {args.input_dir}", code=1)
    if not args.input_dir.is_dir():
        _exit_error(f"Input path is not a directory: {args.input_dir}", code=1)

    # Auto-append .xlsx if missing
    if args.output.suffix.lower() != ".xlsx":
        args.output = args.output.with_suffix(args.output.suffix + ".xlsx")
        log.debug("Appended .xlsx to output path: %s", args.output)

    # Create parent directories if needed
    try:
        args.output.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        _exit_error(f"Cannot create output directory {args.output.parent}: {exc}", code=1)


# ---------------------------------------------------------------------------
# T2 — CSV discovery and loading
# ---------------------------------------------------------------------------

def discover_csvs(input_dir: Path) -> list[Path]:
    """
    Return a sorted list of .csv files in input_dir (non-recursive, case-insensitive).

    Args:
        input_dir: Directory to scan.

    Returns:
        Sorted list of Path objects for .csv files found.
    """
    csvs = sorted(
        p for p in input_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".csv"
    )
    log.debug("Discovered %d CSV file(s) in %s", len(csvs), input_dir)
    return csvs


def load_csv(path: Path) -> tuple[Optional[pd.DataFrame], Optional[str]]:
    """
    Load a CSV file into a DataFrame, with encoding fallback.

    Attempts utf-8-sig (handles BOM) first, then latin-1.

    Args:
        path: Path to the CSV file.

    Returns:
        (DataFrame, None) on success, or (None, error_message) on failure.
    """
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            df = pd.read_csv(path, encoding=encoding, low_memory=False)
            log.debug("Loaded %s (%d rows, %d cols) with encoding=%s",
                      path.name, len(df), len(df.columns), encoding)
            return df, None
        except pd.errors.EmptyDataError:
            return None, "empty file (no data)"
        except UnicodeDecodeError:
            log.debug("Encoding %s failed for %s, trying next.", encoding, path.name)
            continue
        except Exception as exc:  # noqa: BLE001
            return None, str(exc)

    return None, "Could not decode file (tried utf-8-sig, latin-1)"


# ---------------------------------------------------------------------------
# T3 — Sheet name sanitisation and deduplication
# ---------------------------------------------------------------------------

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def sanitise_sheet_name(stem: str, used_names: set[str]) -> tuple[str, set[str]]:
    """
    Produce a valid, unique Excel sheet name from a file stem.

    Rules applied:
    - Strip/replace invalid characters ( [ ] : * ? / \\ ) with underscore.
    - Truncate to SHEET_NAME_MAX_LEN (31).
    - Rename if the result is "Summary" (reserved for the summary sheet).
    - Deduplicate against used_names using _2, _3, ... suffixes.

    Args:
        stem:       The CSV filename without extension.
        used_names: Set of already-assigned sheet names (modified in place).

    Returns:
        (sanitised_name, updated_used_names).
    """
    name = _INVALID_SHEET_CHARS.sub("_", stem).strip()
    if not name:
        name = "Sheet"

    # Reserve "Summary" for the summary sheet
    if name.lower() == RESERVED_SUMMARY_STEM:
        name = "Summary_data"

    name = name[:SHEET_NAME_MAX_LEN]

    # Deduplicate
    candidate = name
    counter = 2
    while candidate.lower() in {n.lower() for n in used_names}:
        suffix = f"_{counter}"
        candidate = name[: SHEET_NAME_MAX_LEN - len(suffix)] + suffix
        counter += 1

    used_names.add(candidate)
    log.debug("Sheet name: %r -> %r", stem, candidate)
    return candidate, used_names


# ---------------------------------------------------------------------------
# T4 — Data sheet writing helpers
# ---------------------------------------------------------------------------

def _auto_col_widths(ws: Worksheet, df: pd.DataFrame) -> None:
    """
    Set column widths based on the max character length in headers and a data sample.

    Args:
        ws:  The openpyxl worksheet to size.
        df:  The DataFrame whose data was written (used for sampling).
    """
    sample = df.head(200)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(col_name))
        try:
            data_max = sample.iloc[:, col_idx - 1].astype(str).str.len().max()
            data_max = int(data_max) if not pd.isna(data_max) else 0
        except Exception:  # noqa: BLE001
            data_max = 0
        width = min(max(header_len, data_max, COL_WIDTH_MIN), COL_WIDTH_MAX)
        ws.column_dimensions[col_letter].width = width + 2  # padding


def write_data_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Write a DataFrame to a new worksheet with styled headers and auto-sized columns.

    Truncates data to Excel limits if necessary and logs a warning.

    Args:
        wb:          The openpyxl Workbook to add the sheet to.
        sheet_name:  Name for the new worksheet.
        df:          DataFrame to write.
    """
    # Check and enforce Excel limits
    original_rows, original_cols = df.shape
    if original_rows > EXCEL_MAX_ROWS - 1:  # -1 for header
        log.warning(
            "Sheet %r: truncating %d rows to Excel limit of %d data rows.",
            sheet_name, original_rows, EXCEL_MAX_ROWS - 1,
        )
        df = df.iloc[: EXCEL_MAX_ROWS - 1]
    if original_cols > EXCEL_MAX_COLS:
        log.warning(
            "Sheet %r: truncating %d columns to Excel limit of %d columns.",
            sheet_name, original_cols, EXCEL_MAX_COLS,
        )
        df = df.iloc[:, :EXCEL_MAX_COLS]

    ws = wb.create_sheet(title=sheet_name)

    # Write header row with styling
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Replace NaN with None so openpyxl writes blank cells instead of "nan"
    df = df.where(pd.notna(df), other=None)

    # Write data rows
    for row_tuple in df.itertuples(index=False, name=None):
        ws.append(list(row_tuple))

    # Auto-size columns
    _auto_col_widths(ws, df)

    log.debug("Wrote sheet %r: %d rows, %d columns.", sheet_name, len(df), len(df.columns))


# ---------------------------------------------------------------------------
# T5 — Summary sheet writing
# ---------------------------------------------------------------------------

def write_summary_sheet(
    wb: Workbook,
    results: list[ResultEntry],
) -> int:
    """
    Create and populate the Summary sheet at position 0.

    The sheet contains: File, Rows, Columns, Column Names — one row per CSV.
    Failed CSVs appear with Rows = "ERROR" and Column Names = error message.

    Args:
        wb:      The openpyxl Workbook (summary sheet will be inserted at index 0).
        results: List of dicts with keys: file, rows, columns, column_names, error.

    Returns:
        The row index of the last data row (1-based), for chart anchoring.
    """
    ws = wb.create_sheet(title=SUMMARY_SHEET_NAME, index=0)

    # Header row
    headers = ["File", "Rows", "Columns", "Column Names"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    ws.freeze_panes = "A2"

    # Data rows
    for r in results:
        if r["error"]:
            ws.append([r["file"], "ERROR", "ERROR", r["error"]])
        else:
            ws.append([r["file"], r["rows"], r["columns"], r["column_names"]])

    # Style column-names column for text wrapping
    last_row = len(results) + 1  # 1-indexed; row 1 = header
    for row_idx in range(2, last_row + 1):
        cell = ws.cell(row=row_idx, column=4)
        cell.alignment = WRAP_ALIGN
        ws.row_dimensions[row_idx].height = 30

    # Auto-size first three columns; cap column-names column
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = COL_WIDTH_MAX

    return last_row


# ---------------------------------------------------------------------------
# T6 — Bar chart creation
# ---------------------------------------------------------------------------

def add_row_count_chart(ws: Worksheet, last_data_row: int, num_rows_with_data: int) -> None:
    """
    Add a vertical bar chart to the Summary sheet showing row counts per file.

    Chart is anchored two rows below the last data row.
    ERROR rows are included in the category axis but excluded from the numeric series
    by leaving those cells blank (they were written as the string "ERROR", which
    openpyxl/Excel treats as a gap in a numeric series).

    Args:
        ws:                  The Summary worksheet.
        last_data_row:       1-based index of the last data row (not counting header).
        num_rows_with_data:  Number of result rows written (excluding header).
    """
    if num_rows_with_data == 0:
        log.debug("No data rows; skipping chart creation.")
        return

    chart = BarChart()
    chart.type = "col"
    chart.title = "Row Counts per File"
    chart.y_axis.title = "Row Count"
    chart.x_axis.title = "File"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    chart.grouping = "clustered"

    # Column B (Rows) is numeric data; column A (File) is categories.
    # last_data_row is already the 1-based sheet row of the final result entry
    # (row 1 = header, row 2 = first result, ..., row last_data_row = last result).
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=last_data_row)
    categories_ref = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    # Anchor two rows below the data table
    anchor_row = last_data_row + 3  # +1 for header, +2 for gap
    ws.add_chart(chart, f"A{anchor_row}")

    log.debug("Bar chart added, anchored at row %d.", anchor_row)


# ---------------------------------------------------------------------------
# T7 — Workbook save
# ---------------------------------------------------------------------------

def save_workbook(wb: Workbook, output_path: Path) -> None:
    """
    Save the workbook to disk, with an actionable message for PermissionError.

    Args:
        wb:          The Workbook to save.
        output_path: Destination file path.

    Raises:
        SystemExit: On PermissionError or other OSError.
    """
    try:
        wb.save(output_path)
        log.info("Workbook saved: %s", output_path.resolve())
    except PermissionError:
        _exit_error(
            f"Cannot write to {output_path} — the file may be open in Excel. "
            "Close it and retry.",
            code=2,
        )
    except OSError as exc:
        _exit_error(f"Failed to save workbook: {exc}", code=2)


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _exit_error(message: str, code: int = 2) -> None:
    """
    Log an error and exit with the given code.

    Args:
        message: Human-readable error message.
        code:    Exit code (1 = user error, 2 = runtime error).
    """
    log.error(message)
    sys.exit(code)


def _setup_logging(verbose: bool) -> None:
    """
    Configure root logger to emit to stderr.

    Args:
        verbose: If True, set level to DEBUG; otherwise INFO.
    """
    level = logging.DEBUG if verbose else logging.INFO
    handler = logging.StreamHandler(sys.stderr)
    handler.setFormatter(
        logging.Formatter("%(levelname)s: %(message)s")
    )
    logging.basicConfig(level=level, handlers=[handler], force=True)


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

def run(input_dir: Path, output_path: Path) -> None:
    """
    Main processing pipeline: discover CSVs, write sheets, build summary, save.

    Args:
        input_dir:   Directory containing CSV files.
        output_path: Destination Excel file path.
    """
    # --- Discover CSVs ---
    csv_paths = discover_csvs(input_dir)
    if not csv_paths:
        _exit_error(
            f"No CSV files found in {input_dir}. Nothing to convert.",
            code=1,
        )

    log.info("Found %d CSV file(s) in %s", len(csv_paths), input_dir)

    # --- Load CSVs and assign sheet names ---
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default empty sheet

    used_names: set[str] = {SUMMARY_SHEET_NAME}  # Reserve summary name upfront
    results: list[ResultEntry] = []
    successful_count = 0
    failed_count = 0

    for path in csv_paths:
        df, error = load_csv(path)

        if df is None:
            log.warning("Skipping %s: %s", path.name, error)
            results.append({
                "file": path.name,
                "rows": None,
                "columns": None,
                "column_names": error,
                "error": error,
            })
            failed_count += 1
            continue

        if len(df) > LARGE_FILE_WARN_ROWS:
            log.info(
                "%s has %d rows — this may take a moment to write.",
                path.name, len(df),
            )

        sheet_name, used_names = sanitise_sheet_name(path.stem, used_names)
        write_data_sheet(wb, sheet_name, df)

        col_names = ", ".join(str(c) for c in df.columns)
        results.append({
            "file": path.name,
            "rows": len(df),
            "columns": len(df.columns),
            "column_names": col_names,
            "error": None,
        })
        successful_count += 1

    if successful_count == 0:
        _exit_error(
            "No CSV files could be read successfully. Check warnings above.",
            code=1,
        )

    # --- Write summary sheet ---
    last_data_row = write_summary_sheet(wb, results)

    # --- Add bar chart ---
    summary_ws = wb[SUMMARY_SHEET_NAME]
    add_row_count_chart(summary_ws, last_data_row, len(results))

    # --- Save ---
    save_workbook(wb, output_path)

    log.info(
        "Done. %d sheet(s) written, %d file(s) skipped. Output: %s",
        successful_count,
        failed_count,
        output_path,
    )


def main() -> None:
    """CLI entry point: parse arguments and invoke the processing pipeline."""
    parser = build_parser()
    args = parser.parse_args()

    _setup_logging(args.verbose)
    validate_args(args)

    log.debug("Input directory: %s", args.input_dir)
    log.debug("Output path:     %s", args.output)

    run(args.input_dir, args.output)


if __name__ == "__main__":
    main()
