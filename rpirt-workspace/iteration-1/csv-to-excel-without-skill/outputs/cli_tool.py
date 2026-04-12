"""
csv_to_excel.py — Convert a folder of CSV files into a formatted Excel workbook.

Produces:
  - One sheet per CSV (sheet name = filename stem, truncated to 31 chars)
  - A "Summary" sheet with row counts and column names per file
  - A bar chart on the Summary sheet visualising row counts

Usage:
    python cli_tool.py --input <folder> --output <file.xlsx> [--encoding utf-8]

Author: generated 2026-04-12
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(fgColor="1F3864", fill_type="solid")          # dark navy
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
SUMMARY_FILL = PatternFill(fgColor="2E75B6", fill_type="solid")         # blue
SUMMARY_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
ALT_ROW_FILL = PatternFill(fgColor="EBF0F8", fill_type="solid")         # light blue tint
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(stem: str, existing: set[str]) -> str:
    """
    Return a unique, Excel-legal sheet name (≤31 chars).

    Illegal chars ``[ ] : * ? / \\`` are replaced with ``_``.
    If the truncated name collides with an existing one, a numeric
    suffix (``_2``, ``_3`` …) is appended.
    """
    for ch in r"[]:*?/\\":
        stem = stem.replace(ch, "_")
    name = stem[:31]
    if name not in existing:
        return name
    for i in range(2, 10_000):
        candidate = f"{stem[:28]}_{i}"
        if candidate not in existing:
            return candidate
    raise RuntimeError(f"Cannot generate unique sheet name for '{stem}'")


def _apply_header_style(ws, row: int, n_cols: int, fill: PatternFill, font: Font) -> None:
    """Bold, coloured header row with borders."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_columns(ws, min_width: int = 8, max_width: int = 45) -> None:
    """Heuristic column widths based on content length."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _freeze_header(ws, row: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# ---------------------------------------------------------------------------
# Core: write one data sheet
# ---------------------------------------------------------------------------

def _write_data_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str, csv_path: Path) -> None:
    """Write *df* to a new sheet with styled headers and alternating row fills."""
    ws = wb.create_sheet(title=sheet_name)

    # Title row
    ws.append([f"Source: {csv_path.name}"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Blank separator
    ws.append([])

    # Data (headers + rows) starting at row 3
    header_row = 3
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        actual_row = header_row + r_idx
        if r_idx == 0:
            _apply_header_style(ws, actual_row, len(row), HEADER_FILL, HEADER_FONT)
        else:
            fill = ALT_ROW_FILL if r_idx % 2 == 0 else PatternFill()
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=actual_row, column=col)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if fill.fill_type:
                    cell.fill = fill

    _autofit_columns(ws)
    _freeze_header(ws, row=header_row)

    # Row count info in status bar area (bottom-left of data, one row below)
    last_data_row = header_row + len(df)
    note_cell = ws.cell(row=last_data_row + 2, column=1)
    note_cell.value = f"{len(df):,} rows  ·  {len(df.columns)} columns"
    note_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")


# ---------------------------------------------------------------------------
# Core: write summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(
    wb: Workbook,
    records: list[dict],
) -> None:
    """
    Write a Summary sheet with:
      - A table: File | Sheet | Rows | Columns | Column Names
      - A bar chart: Rows per file
    """
    ws = wb.create_sheet(title="Summary", index=0)

    # ---- Title ----
    ws.append(["CSV → Excel Summary"])
    ws.cell(1, 1).font = TITLE_FONT
    ws.row_dimensions[1].height = 22
    ws.append([])

    # ---- Table header ----
    header_row_idx = 3
    headers = ["File", "Sheet Name", "Row Count", "Column Count", "Column Names"]
    ws.append(headers)
    _apply_header_style(ws, header_row_idx, len(headers), SUMMARY_FILL, SUMMARY_FONT)
    ws.row_dimensions[header_row_idx].height = 18

    # ---- Data rows ----
    for i, rec in enumerate(records):
        data_row = header_row_idx + 1 + i
        ws.append([
            rec["filename"],
            rec["sheet_name"],
            rec["row_count"],
            rec["col_count"],
            ", ".join(rec["columns"]),
        ])
        fill = ALT_ROW_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))
            if fill.fill_type:
                cell.fill = fill
        ws.row_dimensions[data_row].height = 15

    _autofit_columns(ws)
    ws.column_dimensions[get_column_letter(5)].width = 60  # Column Names wider
    _freeze_header(ws, row=header_row_idx)

    # ---- Bar chart ----
    n = len(records)
    if n == 0:
        return

    # Row count values: column C (index 3), rows header+1 … header+n
    data_ref = Reference(
        ws,
        min_col=3,
        max_col=3,
        min_row=header_row_idx,          # include header as series title
        max_row=header_row_idx + n,
    )
    # File name labels: column A, rows data rows only
    cats_ref = Reference(
        ws,
        min_col=1,
        max_col=1,
        min_row=header_row_idx + 1,
        max_row=header_row_idx + n,
    )

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Row Count per CSV File"
    chart.y_axis.title = "Rows"
    chart.x_axis.title = "File"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Data labels
    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal = True
    chart.series[0].dLbls.showLegendKey = False
    chart.series[0].dLbls.showCatName = False
    chart.series[0].dLbls.showSerName = False

    # Anchor chart below the table (2 rows gap)
    chart_anchor_row = header_row_idx + n + 3
    ws.add_chart(chart, f"A{chart_anchor_row}")


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def convert(
    input_folder: Path,
    output_path: Path,
    encoding: str = "utf-8",
    glob: str = "*.csv",
) -> int:
    """
    Convert all CSV files in *input_folder* to a single Excel workbook.

    Parameters
    ----------
    input_folder:
        Directory to scan for CSV files (non-recursive).
    output_path:
        Destination ``.xlsx`` file path.
    encoding:
        Character encoding used when reading CSVs.
    glob:
        File pattern; defaults to ``*.csv``.

    Returns
    -------
    int
        Number of CSV files successfully converted.
    """
    csv_files = sorted(input_folder.glob(glob))
    if not csv_files:
        log.warning("No CSV files found in '%s' matching '%s'.", input_folder, glob)
        return 0

    log.info("Found %d CSV file(s) in '%s'.", len(csv_files), input_folder)

    wb = Workbook()
    # Remove default empty sheet; Summary will be inserted at index 0
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    records: list[dict] = []
    used_names: set[str] = {"Summary"}  # reserve

    for csv_path in csv_files:
        log.info("  Reading: %s", csv_path.name)
        try:
            df = pd.read_csv(
                csv_path,
                encoding=encoding,
                encoding_errors="replace",
                low_memory=False,
            )
        except Exception as exc:
            log.error("  SKIP — could not read '%s': %s", csv_path.name, exc)
            continue

        sheet_name = _safe_sheet_name(csv_path.stem, used_names)
        used_names.add(sheet_name)

        _write_data_sheet(wb, df, sheet_name, csv_path)

        records.append({
            "filename": csv_path.name,
            "sheet_name": sheet_name,
            "row_count": len(df),
            "col_count": len(df.columns),
            "columns": list(df.columns),
        })
        log.info("    → sheet '%s'  (%d rows, %d cols)", sheet_name, len(df), len(df.columns))

    _write_summary_sheet(wb, records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Workbook saved → %s", output_path.resolve())
    return len(records)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="csv-to-excel",
        description="Convert a folder of CSV files into a formatted Excel workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python cli_tool.py --input ./data --output report.xlsx
  python cli_tool.py --input ./data --output report.xlsx --encoding latin-1
  python cli_tool.py --input ./data --output report.xlsx --verbose
""",
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        type=Path,
        metavar="FOLDER",
        help="Folder containing CSV files to convert.",
    )
    parser.add_argument(
        "--output", "-o",
        required=True,
        type=Path,
        metavar="FILE",
        help="Destination Excel file path (e.g. output/report.xlsx).",
    )
    parser.add_argument(
        "--encoding", "-e",
        default="utf-8",
        metavar="ENCODING",
        help="Character encoding of the CSV files (default: utf-8).",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable DEBUG-level logging.",
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser = _build_parser()
    args = parser.parse_args(argv)

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    input_folder: Path = args.input
    if not input_folder.exists():
        log.error("Input folder not found: %s", input_folder)
        sys.exit(1)
    if not input_folder.is_dir():
        log.error("--input must be a directory, got: %s", input_folder)
        sys.exit(1)

    output_path: Path = args.output
    if output_path.suffix.lower() != ".xlsx":
        log.warning("Output path does not end in .xlsx; appending extension.")
        output_path = output_path.with_suffix(".xlsx")

    count = convert(input_folder, output_path, encoding=args.encoding)

    if count == 0:
        log.error("No sheets written — aborting.")
        sys.exit(1)

    print(f"\nDone. {count} sheet(s) written to: {output_path.resolve()}")


if __name__ == "__main__":
    main()
